# -*- coding: utf-8 -*-
import bs4
import requests
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font


def _save_in_xls(_list_news):
    wb = Workbook()
    ws = wb.create_sheet("News")
    row = 1

    cell = ws.cell(row=row, column=1)
    cell.value = 'url'
    cell.font = Font(color="FF0000", bold=True, size=10)

    cell = ws.cell(row=row, column=2)
    cell.value = 'title'
    cell.font = Font(color="FF0000", bold=True, size=10)

    cell = ws.cell(row=row, column=3)
    cell.value = 'time'
    cell.font = Font(color="FF0000", bold=True, size=10)

    cell = ws.cell(row=row, column=4)
    cell.value = 'author'
    cell.font = Font(color="FF0000", bold=True, size=10)

    row += 1
    for obj in _list_news:
        cell = ws.cell(row=row, column=1)
        cell.value = obj['url']

        cell = ws.cell(row=row, column=2)
        cell.value = obj['title']

        cell = ws.cell(row=row, column=3)
        cell.value = obj['time']

        cell = ws.cell(row=row, column=4)
        cell.value = obj['author']

        row += 1

    wb.save('news.xlsx')


response = requests.get('https://itc.ua/')


instance_page = bs4.BeautifulSoup(response.text, "html.parser")

content = instance_page.find('main', {'id': 'content'})
rows = content.find_all('div', {'class': 'row'})

_urls = []
_list_news = []
for row in rows:

    h2_tag = row.find('h2')
    if not h2_tag:
        continue

    tag_a = h2_tag.find('a')
    url = tag_a.get('href')

    if not url:
        continue

    if url in _urls:
        continue

    title = tag_a.get_text()

    info_tag = row.find('div', {'class': 'entry-header'})
    time = info_tag.find('time').get('datetime')
    author = info_tag.find('a').get_text()

    response = requests.get(url)
    instance_page = bs4.BeautifulSoup(response.text, "html.parser")

    content_text = instance_page.find('div', {"class": 'post-txt'})
    text = content_text.find_all('p')
    _text = ""
    for obj_p in text:
        _text += obj_p.get_text()

    _urls.append(url)
    _article = {
        'url': url,
        'title': title,
        'time': time,
        'author': author,
        '_text': _text
    }
    _list_news.append(_article)


_save_in_xls(_list_news)


# Load Xlsx
from openpyxl import load_workbook
# wb = load_workbook(filename='news.xlsx')
# ws = wb['News']
#
#
# max_row = ws.max_row
#
# for row in range(2, max_row):
#     url = ws.cell(row=row, column=1).value
#
#     title = ws.cell(row=row, column=2).value
#
#     print(title)

# Create formula
# wb = Workbook()
# ws = wb.active
#
# ws.cell(row=1, column=1).value = 1
# ws.cell(row=1, column=2).value = 2
# ws.cell(row=1, column=3).value = "=A1+B1"
#
# wb.save('formulas.xlsx')

# Read formula
# from openpyxl import load_workbook
#
# wb = load_workbook(filename='formulas.xlsx')
#
# ws = wb.active
#
# value = ws.cell(row=1, column=3).value
# print(value)